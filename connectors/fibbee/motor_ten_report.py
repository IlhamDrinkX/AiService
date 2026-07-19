"""
Отчёт "токи моторов и температура ТЭН" по кофейным комплексам Fibbee ERP —
по запросу пользователя (не входит в sync.py, генерируется отдельно).

Источник — то же локальное хранилище, что заполняет sync.py
(FIBBEE_DB_PATH / DB_PATH, таблица fibbee_orders, колонка product_dump).
Никаких новых запросов к API не делает — только читает уже
синхронизированные данные, поэтому перед первым запуском нужен свежий
python sync.py.

Что внутри productDump (см. подробности и обоснование порогов в листе
"Методика и ограничения" самого отчёта, а также в connectors/fibbee/README.md):

- Часть заказов ("drinkx"-формат — фирменное оборудование DrinkX) содержит
  вложенный массив partResults — по записи на каждый мотор-модуль
  (coffee/milk/water), с полем avgCurrent (средний ток модуля, А). Это и
  есть "токи моторов при приготовлении" — статус мотора считается по
  Z-отклонению тока события от среднего по своей группе (комплекс+модуль).
- Часть заказов (другой формат — сторонняя кофемашина "eversys", пока
  встречалось только на одном комплексе) содержит boilerTemp/waterTemp —
  по ним отдельно считается статус ТЭН.
- Тока ТЭНа (в амперах) в данных нет ни у одного формата — только
  температура бойлера/воды, статус ТЭН оценивается по ней.

Использование:
    cd connectors/fibbee
    python motor_ten_report.py
    # результат: ./reports/fibbee_motor_ten_report_<YYYYMMDD_HHMM>.xlsx

Пороги эвристические (3 стандартных отклонения по току, 85-115°C и ±4°C по
температуре бойлера) — см. лист "Методика и ограничения" за полным
обоснованием и как их пересмотреть.
"""

import json
import os
import sqlite3
import statistics
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

DB_PATH = os.environ.get("DB_PATH", "./data/fibbee.db")
REPORTS_DIR = os.environ.get("FIBBEE_REPORTS_DIR", "./reports")

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FILL2 = PatternFill("solid", fgColor="548235")
BASE_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# пороги — см. лист "Методика и ограничения" за обоснованием
Z_SCORE_THRESHOLD = 3.0
ZERO_CURRENT_THRESHOLD = 0.02
BOILER_TEMP_ABS_MIN, BOILER_TEMP_ABS_MAX = 85, 115
BOILER_TEMP_DEVIATION_THRESHOLD = 4


def style_header(ws, row, headers, fill):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 34


def style_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BASE_FONT
        cell.border = BORDER


def load_data(db_path: str) -> dict:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    rows = conn.execute("""
        SELECT o.sales_point_id, sp.name_ru, o.order_id, o.number, o.status, o.received_at, o.product_dump
        FROM fibbee_orders o LEFT JOIN fibbee_sales_points sp ON sp.sales_point_id = o.sales_point_id
        WHERE o.product_dump LIKE '%partResults%'
    """).fetchall()

    motor_events = []
    for r in rows:
        pd = json.loads(r["product_dump"])
        complex_name = r["name_ru"] or r["sales_point_id"]
        for part in pd.get("partResults") or []:
            if part.get("avgCurrent") is None:
                continue
            motor_events.append({
                "complex": complex_name,
                "type": part.get("type") or "?",
                "order_number": r["number"],
                "order_id": r["order_id"],
                "received_at": r["received_at"],
                "avgCurrent": part.get("avgCurrent"),
                "peakTemp": part.get("peakTemp"),
                "avgTemp": part.get("avgTemp"),
                "pumpPower": part.get("pumpPower"),
                "totalTime": part.get("totalTime"),
            })
    motor_events.sort(key=lambda x: (x["complex"], x["type"], x["received_at"]))

    rows2 = conn.execute("""
        SELECT o.sales_point_id, sp.name_ru, o.order_id, o.number, o.status, o.received_at, o.product_dump
        FROM fibbee_orders o LEFT JOIN fibbee_sales_points sp ON sp.sales_point_id = o.sales_point_id
        WHERE o.product_dump LIKE '%boilerTemp%'
        ORDER BY sp.name_ru, o.received_at
    """).fetchall()
    temp_orders = []
    for r in rows2:
        pd = json.loads(r["product_dump"])
        temp_orders.append({
            "complex": r["name_ru"] or r["sales_point_id"],
            "order_number": r["number"], "order_id": r["order_id"], "status": r["status"],
            "received_at": r["received_at"],
            "boilerTemp": pd.get("boilerTemp"), "waterTemp": pd.get("waterTemp"),
            "milkTemp": pd.get("milkTemp") or None, "steamPress": pd.get("steamPress"),
        })

    total_orders = conn.execute("SELECT COUNT(*) FROM fibbee_orders").fetchone()[0]
    total_complexes = conn.execute("SELECT COUNT(*) FROM fibbee_sales_points").fetchone()[0]
    conn.close()

    return {
        "motor_events": motor_events,
        "temp_orders": temp_orders,
        "total_orders_in_window": total_orders,
        "total_complexes_synced": total_complexes,
        "motor_complexes": sorted(set(e["complex"] for e in motor_events)),
        "temp_complexes": sorted(set(o["complex"] for o in temp_orders)),
    }


def build_workbook(data: dict) -> Workbook:
    motor_events = data["motor_events"]
    temp_orders = data["temp_orders"]
    total_orders_window = data["total_orders_in_window"]
    total_complexes = data["total_complexes_synced"]
    motor_complexes = data["motor_complexes"]
    temp_complexes = data["temp_complexes"]

    groups = {}
    for e in motor_events:
        groups.setdefault((e["complex"], e["type"]), []).append(e["avgCurrent"])
    group_stats = {
        key: (statistics.mean(vals), statistics.stdev(vals) if len(vals) > 1 else 0.0)
        for key, vals in groups.items()
    }

    wb = Workbook()

    # ---------------------------------------------------------- Токи моторов
    # Z-отклонение и статус посчитаны один раз в Python (не формулами Excel в
    # каждой строке) — на большой выборке (тысячи строк) формулы на каждую
    # строку упирались в таймаут headless-пересчёта LibreOffice в песочнице
    # разработки. Средний ток/стдев группы, от которых зависит расчёт, —
    # живые формулы на листе "Сводка токов" (см. ниже); при обновлении
    # исходных данных пересчитайте их и перегенерируйте отчёт этим скриптом.
    ws = wb.active
    ws.title = "Токи моторов"
    headers = ["Комплекс", "Модуль (мотор)", "№ заказа", "Order ID", "Получен (UTC)",
               "Ток, A (avgCurrent)", "Пиковая T, °C", "Средняя T, °C", "Мощность насоса", "Время, мс",
               "Ср. ток группы, A", "Стдев тока группы", "Z-отклонение", "Статус мотора"]
    style_header(ws, 1, headers, HEADER_FILL)

    first_row = 2
    group_ranges = {}
    for i, e in enumerate(motor_events):
        r = first_row + i
        key = (e["complex"], e["type"])
        group_ranges.setdefault(key, [r, r])
        group_ranges[key][1] = r

    for i, e in enumerate(motor_events):
        r = first_row + i
        key = (e["complex"], e["type"])
        m, s = group_stats[key]
        cur = e["avgCurrent"]
        z = (cur - m) / s if s > 0 else 0.0
        if cur is not None and cur <= ZERO_CURRENT_THRESHOLD:
            status = "Проверить (нулевой ток)"
        elif abs(z) > Z_SCORE_THRESHOLD:
            status = "Проверить (аномальный ток)"
        else:
            status = "ОК"
        ws.cell(row=r, column=1, value=e["complex"])
        ws.cell(row=r, column=2, value=e["type"])
        ws.cell(row=r, column=3, value=e["order_number"])
        ws.cell(row=r, column=4, value=e["order_id"])
        ws.cell(row=r, column=5, value=e["received_at"])
        ws.cell(row=r, column=6, value=round(cur, 4) if cur is not None else None)
        ws.cell(row=r, column=7, value=round(e["peakTemp"], 2) if e["peakTemp"] is not None else None)
        ws.cell(row=r, column=8, value=round(e["avgTemp"], 2) if e["avgTemp"] is not None else None)
        ws.cell(row=r, column=9, value=e["pumpPower"])
        ws.cell(row=r, column=10, value=e["totalTime"])
        ws.cell(row=r, column=11, value=round(m, 4))
        ws.cell(row=r, column=12, value=round(s, 4))
        ws.cell(row=r, column=13, value=round(z, 3))
        ws.cell(row=r, column=14, value=status)
        style_row(ws, r, 14)

    for i, w in enumerate([22, 12, 9, 22, 20, 15, 12, 13, 13, 10, 14, 14, 13, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---------------------------------------------------------- Сводка токов
    ws2 = wb.create_sheet("Сводка токов")
    headers2 = ["Комплекс", "Модуль", "Средний ток, A", "Стдев тока", "Мин ток, A", "Макс ток, A",
                "Событий", "Заказов-выбросов", "Заказов с нулевым током"]
    style_header(ws2, 1, headers2, HEADER_FILL2)

    for i, (key, (r_start, r_end)) in enumerate(group_ranges.items()):
        cx, tp = key
        r = 2 + i
        ws2.cell(row=r, column=1, value=cx)
        ws2.cell(row=r, column=2, value=tp)
        ws2.cell(row=r, column=3, value=f"=AVERAGE('Токи моторов'!$F${r_start}:$F${r_end})")
        ws2.cell(row=r, column=4,
                 value=f"=IF(COUNT('Токи моторов'!$F${r_start}:$F${r_end})<2,0,STDEV('Токи моторов'!$F${r_start}:$F${r_end}))")
        ws2.cell(row=r, column=5, value=f"=MIN('Токи моторов'!$F${r_start}:$F${r_end})")
        ws2.cell(row=r, column=6, value=f"=MAX('Токи моторов'!$F${r_start}:$F${r_end})")
        ws2.cell(row=r, column=7, value=f"=COUNTA('Токи моторов'!$F${r_start}:$F${r_end})")
        ws2.cell(row=r, column=8,
                 value=f'=COUNTIFS(\'Токи моторов\'!$A${r_start}:$A${r_end},A{r},\'Токи моторов\'!$B${r_start}:$B${r_end},B{r},\'Токи моторов\'!$N${r_start}:$N${r_end},"Проверить (аномальный ток)")')
        ws2.cell(row=r, column=9,
                 value=f'=COUNTIFS(\'Токи моторов\'!$A${r_start}:$A${r_end},A{r},\'Токи моторов\'!$B${r_start}:$B${r_end},B{r},\'Токи моторов\'!$N${r_start}:$N${r_end},"Проверить (нулевой ток)")')
        style_row(ws2, r, 9)

    for i, w in enumerate([22, 10, 15, 12, 11, 11, 10, 16, 20], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ---------------------------------------------------------- Температура бойлера
    ws3 = wb.create_sheet("Температура бойлера")
    headers3 = ["Комплекс", "№ заказа", "Order ID", "Статус заказа", "Получен (UTC)",
                "Т бойлера, °C", "Т воды, °C", "Т молока, °C", "Давление пара",
                "Откл. Т бойлера от среднего", "Статус ТЭН"]
    style_header(ws3, 1, headers3, HEADER_FILL)

    t_first = 2
    t_last = t_first + len(temp_orders) - 1
    for i, o in enumerate(temp_orders):
        r = t_first + i
        ws3.cell(row=r, column=1, value=o["complex"])
        ws3.cell(row=r, column=2, value=o["order_number"])
        ws3.cell(row=r, column=3, value=o["order_id"])
        ws3.cell(row=r, column=4, value=o["status"])
        ws3.cell(row=r, column=5, value=o["received_at"])
        ws3.cell(row=r, column=6, value=o["boilerTemp"])
        ws3.cell(row=r, column=7, value=o["waterTemp"])
        ws3.cell(row=r, column=8, value=o["milkTemp"])
        ws3.cell(row=r, column=9, value=o["steamPress"])
        if t_last >= t_first:
            ws3.cell(row=r, column=10, value=f"=F{r}-AVERAGE($F${t_first}:$F${t_last})")
            ws3.cell(row=r, column=11,
                     value=f'=IF(OR(F{r}<{BOILER_TEMP_ABS_MIN},F{r}>{BOILER_TEMP_ABS_MAX},ABS(J{r})>{BOILER_TEMP_DEVIATION_THRESHOLD}),"Проверить","ОК")')
        style_row(ws3, r, 11)

    for i, w in enumerate([22, 9, 22, 13, 20, 13, 11, 12, 12, 15, 12], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ---------------------------------------------------------- Методика
    ws4 = wb.create_sheet("Методика и ограничения")
    ws4.column_dimensions["A"].width = 105
    lines = [
        ("Отчёт: токи моторов и температура ТЭН по кофейным комплексам Fibbee ERP", TITLE_FONT),
        (f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')} (данные из локальной синхронизации connectors/fibbee, python sync.py)", NOTE_FONT),
        ("", BASE_FONT),
        ("Источник данных", BOLD_FONT),
        ("GET /v1/orders/list (Fibbee ERP), поле productDump каждого заказа. Локально: "
         "connectors/fibbee/data/fibbee.db, таблица fibbee_orders, колонка product_dump.", BASE_FONT),
        (f"Окно данных: последняя синхронизация, всего {total_orders_window} заказов по {total_complexes} комплексам в базе.", BASE_FONT),
        ("", BASE_FONT),
        ("Токи моторов — где нашлись и что это", BOLD_FONT),
        ("У части заказов productDump содержит вложенный массив partResults — по одной записи на "
         "каждый мотор-модуль (coffee/milk/water), участвовавший в приготовлении. Внутри — "
         "avgCurrent (средний ток модуля за время работы, А), а также сырой tempLog с колонками "
         "pump_R_IS/pump_L_IS (мгновенный ток правого/левого насоса) и колонками PID-цикла "
         "(target/measurement/input/output) — в этот отчёт вошли только агрегаты (avgCurrent, "
         "peakTemp, pumpPower); секундный tempLog не разворачивался — при необходимости он есть "
         "в raw_json целиком.", BASE_FONT),
        (f"Формат partResults встретился у {len(motor_complexes)} из {total_complexes} комплексов "
         f"(это фирменное оборудование DrinkX — формат \"drinkx\"): {', '.join(motor_complexes) or '—'}.", BASE_FONT),
        ("Флаг success внутри partResults в имеющихся данных всегда true — по нему отдельные сбои "
         "не отличить, поэтому статус мотора здесь основан на статистике тока, а не на этом флаге.", BASE_FONT),
        ("", BASE_FONT),
        ("Температура бойлера (ТЭН)", BOLD_FONT),
        (f"Отдельно от токов — по температуре бойлера/воды данные нашлись только у: "
         f"{', '.join(temp_complexes) or '—'} (формат productDump \"eversys\", сторонняя "
         "кофемашина). У остальных комплексов, включая те, что с токами моторов выше, полей "
         "boilerTemp/waterTemp в заказе нет вообще — это два РАЗНЫХ формата телеметрии от разных "
         "производителей оборудования, не пробел синхронизации.", BASE_FONT),
        ("", BASE_FONT),
        ("Методика (пороги — статистическая эвристика по имеющимся данным, НЕ паспортные допуски производителя)", BOLD_FONT),
        (f"Статус мотора: события группируются по паре (комплекс, модуль). Z-отклонение — во "
         f"сколько стандартных отклонений ток события отличается от среднего по своей группе. "
         f"«Проверить (аномальный ток)» — если |Z|>{Z_SCORE_THRESHOLD}. «Проверить (нулевой ток)» "
         f"— если ток ≤{ZERO_CURRENT_THRESHOLD}А. Иначе — «ОК».", BASE_FONT),
        (f"Статус ТЭН: «Проверить», если температура бойлера заказа выходит за "
         f"{BOILER_TEMP_ABS_MIN}–{BOILER_TEMP_ABS_MAX}°C ИЛИ отклоняется от среднего по комплексу "
         f"больше чем на {BOILER_TEMP_DEVIATION_THRESHOLD}°C.", BASE_FONT),
        ("Пороги стоит откалибровать по факту, когда накопится история реальных поломок/ремонтов "
         "и будет с чем сверить статистику — сейчас это первая версия эвристики.", BASE_FONT),
        ("", BASE_FONT),
        ("Как обновить отчёт", BOLD_FONT),
        ("1) python sync.py (при необходимости увеличить FIBBEE_SYNC_DAYS в .env). "
         "2) python motor_ten_report.py.", BASE_FONT),
    ]
    for r, (text, font) in enumerate(lines, start=1):
        cell = ws4.cell(row=r, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=False, vertical="top")

    wb._sheets = [wb["Токи моторов"], wb["Сводка токов"], wb["Температура бойлера"], wb["Методика и ограничения"]]
    wb.active = 0
    return wb


def main():
    data = load_data(DB_PATH)
    wb = build_workbook(data)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    out_path = os.path.join(REPORTS_DIR, f"fibbee_motor_ten_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(out_path)
    print(f"Готово: {out_path}")
    print(f"Токов-событий: {len(data['motor_events'])} ({len(data['motor_complexes'])} комплексов)")
    print(f"Заказов с Т бойлера: {len(data['temp_orders'])} ({len(data['temp_complexes'])} комплексов)")


if __name__ == "__main__":
    main()
